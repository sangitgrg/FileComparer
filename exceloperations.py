from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import YELLOW,GREEN
from os import listdir

new_lft_folder = 'C:/USERDATA/Development/LFTComparer/TagLFTAll/New'
old_lft_folder = 'C:/USERDATA/Development/LFTComparer/TagLFTAll/Old'
old_ws_dict = {}

for x in listdir(new_lft_folder):
    print('processing ' + x)
    new_wb = load_workbook(new_lft_folder + '/' + x)
    new_ws = new_wb.worksheets[0]

    old_wb = load_workbook(old_lft_folder + '/' + x, read_only= True)
    old_ws = old_wb.worksheets[0]
    for row in old_ws.iter_rows(min_row=3, max_row=old_ws.max_row,min_col = 2, max_col=2):
        for cell in row:
            old_ws_dict[cell.value] = cell

    for row_cells in new_ws.iter_rows(min_row=3, max_row=new_ws.max_row, max_col=new_ws.max_column):
        print(row_cells[1].value)
        if row_cells[1].value in old_ws_dict:
            for cell in row_cells:
                # print(old_ws.cell(row = cell.row,column = cell.column).value)
                if(cell.value != old_ws.cell(row=cell.row, column=cell.column).value):
                    cell.fill = PatternFill(fgColor=YELLOW, fill_type="solid")
        else:
            row_cells.fill = PatternFill(fgColor= '00008000', fill_type="solid")

def store_old_ws(self, oldworksheet):
    for row in oldworksheet:
        for cell in row:
            old_ws_dict[cell.value] = (cell.row, cell.col_idx)
