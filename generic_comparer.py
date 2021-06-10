import pandas as pd
import numpy as np
import configparser
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from os import listdir, path

# reading from config
_config_object = configparser.ConfigParser()
_config_object.read('config.ini')
_new_excel_file = _config_object['PathConfig']['newFolderPath']
_old_excel_file = _config_object['PathConfig']['oldFolderPath']
_new_excel_sheet_name = _config_object['ExcelConfig']['newExcelsheetName']
_old_excel_sheet_name = _config_object['ExcelConfig']['oldExcelsheetName']
_uniqueKeyForCompare = _config_object['ExcelConfig']['uniqueKey']


class GenericComparer:

    print('processing ' + _new_excel_file)

    def startComparing():

        # seperating filename and extension
        filename = path.splitext(_new_excel_file)[0]

        df_new = pd.read_excel(
            _new_excel_file, sheet_name=_new_excel_sheet_name)
        df_old = pd.read_excel(
            _old_excel_file, sheet_name=_old_excel_sheet_name)

        # dropping duplicate record to get only new record
        df_new_tags = pd.concat([df_new, df_old]).drop_duplicates(
            subset=[_uniqueKeyForCompare], keep=False)

        # deleting new record
        # to make old and new same amount of records
        new_tags_list = df_new_tags.index.values.tolist()
        df_new = df_new.drop(new_tags_list)

        # removing blank rows because after delete there might be null
        df_new = df_new[df_new[_uniqueKeyForCompare].notna()]
        df_old = df_old[df_old[_uniqueKeyForCompare].notna()]  # just incase

        df_new = df_new.reset_index(drop=True)
        df_old = df_old.reset_index(drop=True)

        # getting the difference between two
        # comparing with old file for getting new records
        # use unique column/key for comparing
        df_new['StatusMatch'] = np.where(
            df_new['Status'] == df_old['Status'], 'True', 'False')   # this will add new column 'StatusMatch'
        df_new['DescriptionMatch'] = np.where(
            df_new['Description'] == df_old['Description'], 'True', 'False')  # this will add new column 'DescriptionMatch'

        # actual comparison
        difference = df_new[df_new.values != df_old.values]
        # getting row and col index for differences
        rowCol = np.argwhere(difference.notnull().values).tolist()

        if rowCol:
            new_wb = load_workbook(_new_excel_file)
            new_ws = new_wb.worksheets[0]

        for row in rowCol:
            colLetter = get_column_letter(row[1]+1)
            new_ws[colLetter + str(row[0]+3)].fill = PatternFill(start_color="ededa8", end_color="ededa8",
                                                                 fill_type="solid")
        if rowCol:
            new_wb.save('./' + filename + '_highlighted.xlsx')

        print(filename + ' processing completed')
