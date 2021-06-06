import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from os import listdir, path
import configparser

config_object = configparser.ConfigParser()
config_object.read("config.ini")
path_config = config_object["PathConfig"]

new_lft_folder = path_config['newFolderPath']
old_lft_folder = path_config['oldFolderPath']

new_folder_files = listdir(new_lft_folder)

new_wb=None
wb_dict = {}

def load_workbook_from(wb_file_path):
    #global wb_dict
    filename = path.splitext(path.basename(wb_file_path))[0] # seperating filename and extension
    new_wb = load_workbook(wb_file_path)
    wb_dict[filename] = new_wb
    print(wb_dict.values())
    #return wb_dict

def save_workbook_from():
    print(wb_dict.values())    

def start_compare(filePath,sheetName,wbDict,return_dict):
    print('processing ' + sheetName)
    filename = path.splitext(filePath)[0] # seperating filename and extension
    df_new = pd.read_excel(new_lft_folder + '/' + filePath, sheet_name=sheetName , skiprows=1)
    df_old = pd.read_excel(old_lft_folder + '/' + filePath, sheet_name=sheetName, skiprows=1)

    if 'Tag' in filename:
        if 'Pipeline' in sheetName:
            df_new_tags = pd.concat([df_new, df_old]).drop_duplicates(subset=['Tag No', 'Pipeline Unique ID'], keep=False)
        else:
            if '12.Instrument' in sheetName:
                df_new = df_new[df_new['Tag Type'] != 'Instrument Cable'] # removing instrument cable tags for duplication
                df_old = df_old[df_old['Tag Type'] != 'Instrument Cable'] # removing instrument cable tags for duplication
            df_new_tags = pd.concat([df_new, df_old]).drop_duplicates(subset=['Tag No'], keep=False)
        new_tags_list = df_new_tags.index.values.tolist()
        df_new = df_new.drop(new_tags_list) # deleting new Asset numbers for matching only common
        df_new = df_new[df_new['Tag No'].notna()] # removing blank rows
        df_old = df_old[df_old['Tag No'].notna()]    
    elif 'Asset' in filename:
        df_new_tags = pd.concat([df_new, df_old]).drop_duplicates(subset=['Asset Number'], keep=False)
        new_tags_list = df_new_tags.index.values.tolist() # holding new asset numbers
        df_new = df_new.drop(new_tags_list)
        df_new = df_new[df_new['Asset Number'].notna()] # removing blank rows
        df_old = df_old[df_old['Asset Number'].notna()]
    # getting the difference between two
    difference = df_new[df_new != df_old]
    # getting row and col index for differences
    rowCol = np.argwhere(difference.notnull().values).tolist()

    if rowCol or new_tags_list:
        w_b = wbDict[filename]
        new_ws = w_b[sheetName]
        new_ws.sheet_properties.tabColor = 'FFFF00'
        for row in rowCol:
            colLetter = get_column_letter(row[1]+1)
            new_ws[colLetter + str(row[0]+3)].fill = PatternFill(start_color="ededa8", end_color="ededa8",
                                                                    fill_type="solid")
        if new_tags_list:
            x, maxCol = df_new_tags.shape
            for rows in new_ws.iter_rows(min_row=min(new_tags_list)+3, max_row=max(new_tags_list)+3, min_col=1, max_col=maxCol):
                for cell in rows:
                    cell.fill = PatternFill(start_color="4dab05", end_color="4dab05",
                                            fill_type="solid")
    wb_dict = wbDict        
    return_dict = wbDict
    # if rowCol or new_tags_list:
    #     # new_wb.save('./' + filename + '_highlighted.xlsx')
    #     return_dict[filename] = new_ws
    # else:
    #     return_dict[filename] = "NO_CHANGE"
    print(sheetName +' processing completed')