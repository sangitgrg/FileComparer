import pandas as pd
from multiprocessing import Process, Queue
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
import os
import numpy as np
import multiprocessing
import configparser

config_object = configparser.ConfigParser()
config_object.read("config.ini")
path_config = config_object["PathConfig"]
new_lft_folder = path_config['newFolderPath']
old_lft_folder = path_config['oldFolderPath']
#Q= Queue()

def start_compare(filePath,sheetName,new_ws):
    print('processing ' + sheetName)
    filename = os.path.splitext(filePath)[0] # seperating filename and extension
    df_new = pd.read_excel(new_lft_folder + '/' + filePath, sheet_name=sheetName , skiprows=1)
    df_old = pd.read_excel(old_lft_folder + '/' + filePath, sheet_name=sheetName, skiprows=1)

    if 'Tag' in filename:
        if 'Pipeline' in sheetName:
            df_new_tags = pd.concat([df_new, df_old]).drop_duplicates(subset=['Tag No', 'Pipeline Unique ID'], keep=False)
        else:
            if '12.Instrument' in sheetName:
                df_new = df_new[df_new['Tag Type'] != 'Instrument Cable'] # removing instrument cable tags for duplication
                df_old = df_old[df_old['Tag Type'] != 'Instrument Cable'] # removing instrument cable tags for duplication
            df_new_tags = pd.concat([df_new, df_old]).drop_duplicates(subset=['Tag No'], keep=False)
        new_tags_list = df_new_tags.index.values.tolist()
        df_new = df_new.drop(new_tags_list) # deleting new Asset numbers for matching only common
        df_new = df_new[df_new['Tag No'].notna()] # removing blank rows
        df_old = df_old[df_old['Tag No'].notna()]    
    elif 'Asset' in filename:
        df_new_tags = pd.concat([df_new, df_old]).drop_duplicates(subset=['Asset Number'], keep=False)
        new_tags_list = df_new_tags.index.values.tolist() # holding new asset numbers
        df_new = df_new.drop(new_tags_list)
        df_new = df_new[df_new['Asset Number'].notna()] # removing blank rows
        df_old = df_old[df_old['Asset Number'].notna()]
    # getting the difference between two
    difference = df_new[df_new != df_old]
    # getting row and col index for differences
    rowCol = np.argwhere(difference.notnull().values).tolist()
    #new_ws = new_wb[sheetName]
    if rowCol or new_tags_list:
        new_ws.sheet_properties.tabColor = 'FFFF00'
        for row in rowCol:
            colLetter = get_column_letter(row[1]+1)
            new_ws[colLetter + str(row[0]+3)].fill = PatternFill(start_color="ededa8", end_color="ededa8",
                                                                    fill_type="solid")
        if new_tags_list:
            x, maxCol = df_new_tags.shape
            for rows in new_ws.iter_rows(min_row=min(new_tags_list)+3, max_row=max(new_tags_list)+3, min_col=1, max_col=maxCol):
                for cell in rows:
                    cell.fill = PatternFill(start_color="4dab05", end_color="4dab05",
                                            fill_type="solid")
    #new_wb.save('./' + filename + '_highlighted.xlsx')      
    #Q.put(new_ws)
    print(sheetName +' processing completed')


if __name__ == "__main__":  # confirms that the code is under main function
    try:
        procs = []
        now = datetime.now()
        print('comparison started ' + str(now.hour) + ':' + str(now.minute))
        list_files = os.listdir(new_lft_folder)
        #new_wb = None
        # instantiating process with arguments
        for file_path in list_files:
            print('looking on ' + file_path)
            new_xls = pd.ExcelFile(new_lft_folder + '/' + file_path)
            new_wb = load_workbook(new_lft_folder + '/' + file_path)
            new_sheets = new_xls.sheet_names  # getting list of sheet names
            for sheetName in new_sheets:
                new_ws = new_wb[sheetName]
                proc = Process(target= start_compare,args=(file_path,sheetName,new_ws))
                procs.append(proc)
                proc.start()
            # complete the processes
            for proc in procs:
                #xx = Q.get()
                proc.join()    
            new_wb.save('./' + os.path.splitext(file_path)[0] + '_highlighted.xlsx')      
        
        now = datetime.now()
        print('completed for all files ' + str(now.hour) + ':' + str(now.minute))
        os.system('pause')
    except Exception as e:
        print('oops! error occurred.')
        print(sys.exc_info()[0])
        os.system('pause')

