import pandas as pd
from multiprocessing import Process
from openpyxl import load_workbook
from datetime import datetime
import sys
import os
import multiprocessing
import configparser
import lft_comparer

# this is an attempt to process many number of files at once using python multiprocess
# this is not tested might be bugs

if __name__ == "__main__":  # confirms that the code is under main function
    try:
        config_object = configparser.ConfigParser()
        config_object.read("config.ini")
        path_config = config_object["PathConfig"]

        new_lft_folder = path_config['newFolderPath']
        old_lft_folder = path_config['oldFolderPath']

        list_files = lft_comparer.new_folder_files
        procs = []
        now = datetime.now()
        print('comparison started ' + str(now.hour) + ':' + str(now.minute))
        manager = multiprocessing.Manager()
        return_dict = manager.dict()

        # instantiating process with arguments
        for file_path in list_files:
            print('looking on ' + file_path)
            new_xls = pd.ExcelFile(new_lft_folder + '/' + file_path)
            lft_comparer.load_workbook_from(new_lft_folder + '/' + file_path)
            # new_wb = load_workbook(new_lft_folder + '/' + file_path)
            new_sheets = new_xls.sheet_names  # getting list of sheet names
            for sheetName in new_sheets:
                # new_ws = new_wb[sheetName]
                proc = Process(target=lft_comparer.start_compare, args=(
                    file_path, sheetName, lft_comparer.wb_dict, return_dict))
                procs.append(proc)
                proc.start()

        # complete the processes
        for proc in procs:
            proc.join()

        for x, y in lft_comparer.wb_dict.items():
            y.save('./' + x + '_highlighted.xlsx')

        if return_dict:
            for filename, wb in return_dict.items():
                wb.save('./' + filename + '_highlighted.xlsx')
            print(return_dict.values())
        lft_comparer.save_workbook_from()
        now = datetime.now()
        print('completed for all files ' +
              str(now.hour) + ':' + str(now.minute))
        os.system('pause')
    except Exception as e:
        print('oops! error occurred.')
        print(sys.exc_info()[0])
        os.system('pause')
